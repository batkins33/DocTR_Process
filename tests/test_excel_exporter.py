"""Unit tests for Excel exporter (Issue #12)."""
import tempfile
from datetime import date
from pathlib import Path

import pytest

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.truck_tickets.exporters.excel_exporter import ExcelTrackingExporter


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestExcelTrackingExporter:
    """Test suite for Excel tracking exporter."""

    @pytest.fixture
    def sample_tickets(self):
        """Create sample ticket data for testing."""
        return [
            # Class 2 Contaminated tickets
            {
                "ticket_date": "2024-10-17",
                "ticket_number": "WM-40000001",
                "material": "CLASS_2_CONTAMINATED",
                "source": "PODIUM",
                "destination": "WASTE_MANAGEMENT_LEWISVILLE",
                "ticket_type": "EXPORT",
                "quantity": 15.5,
                "vendor": "WASTE_MANAGEMENT_LEWISVILLE"
            },
            {
                "ticket_date": "2024-10-17",
                "ticket_number": "WM-40000002",
                "material": "CLASS_2_CONTAMINATED",
                "source": "PIER_EX",
                "destination": "WASTE_MANAGEMENT_LEWISVILLE",
                "ticket_type": "EXPORT",
                "quantity": 12.0,
                "vendor": "WASTE_MANAGEMENT_LEWISVILLE"
            },
            # Non-contaminated ticket
            {
                "ticket_date": "2024-10-17",
                "ticket_number": "WM-40000003",
                "material": "NON_CONTAMINATED",
                "source": "SPG",
                "destination": "LDI_YARD",
                "ticket_type": "EXPORT",
                "quantity": 18.0,
                "vendor": "WASTE_MANAGEMENT_LEWISVILLE"
            },
            # Spoils tickets (from other subs' areas)
            {
                "ticket_date": "2024-10-18",
                "ticket_number": "WM-40000004",
                "material": "SPOILS",
                "source": "BECK_SPOILS",
                "destination": "WASTE_MANAGEMENT_LEWISVILLE",
                "ticket_type": "EXPORT",
                "quantity": 14.0,
                "vendor": "WASTE_MANAGEMENT_LEWISVILLE"
            },
            {
                "ticket_date": "2024-10-18",
                "ticket_number": "WM-40000005",
                "material": "SPOILS",
                "source": "NTX_SPOILS",
                "destination": "WASTE_MANAGEMENT_LEWISVILLE",
                "ticket_type": "EXPORT",
                "quantity": 16.0,
                "vendor": "WASTE_MANAGEMENT_LEWISVILLE"
            },
            # Import tickets
            {
                "ticket_date": "2024-10-19",
                "ticket_number": "HD-50000001",
                "material": "FLEXBASE",
                "source": "HEIDELBERG",
                "destination": "PODIUM",
                "ticket_type": "IMPORT",
                "quantity": 20.0,
                "vendor": "HEIDELBERG"
            },
            {
                "ticket_date": "2024-10-19",
                "ticket_number": "HD-50000002",
                "material": "ROCK",
                "source": "HEIDELBERG",
                "destination": "TRACT_2",
                "ticket_type": "IMPORT",
                "quantity": 22.0,
                "vendor": "HEIDELBERG"
            },
        ]

    @pytest.fixture
    def exporter(self):
        """Create exporter instance with test job start date."""
        return ExcelTrackingExporter(job_start_date=date(2024, 7, 1))

    def test_exporter_initialization(self, exporter):
        """Test exporter initializes correctly."""
        assert exporter.job_start_date == date(2024, 7, 1)
        assert exporter.logger is not None

    def test_export_creates_file(self, exporter, sample_tickets):
        """Test that export creates an Excel file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            assert output_path.exists()
            assert output_path.stat().st_size > 0

    def test_export_creates_5_sheets(self, exporter, sample_tickets):
        """Test that export creates exactly 5 sheets."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames
            
            assert len(sheet_names) == 5
            assert "All Daily" in sheet_names
            assert "Class2_Daily" in sheet_names
            assert "Non Contaminated" in sheet_names
            assert "Spoils" in sheet_names
            assert "Import" in sheet_names

    def test_all_daily_sheet_content(self, exporter, sample_tickets):
        """Test All Daily sheet has correct data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb["All Daily"]
            
            # Check headers
            headers = [cell.value for cell in ws[1]]
            assert "Date" in headers
            assert "Day" in headers
            assert "Job Week" in headers
            assert "Job Month" in headers
            assert "Total" in headers
            assert "Class 2" in headers
            assert "Non Contaminated" in headers
            assert "Spoils" in headers
            
            # Check data rows exist
            assert ws.max_row > 1  # More than just header

    def test_class2_daily_sheet_content(self, exporter, sample_tickets):
        """Test Class2_Daily sheet has correct data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Class2_Daily"]
            
            # Check headers include source locations
            headers = [cell.value for cell in ws[1]]
            assert "PODIUM" in headers
            assert "PierEx" in headers
            assert "Total" in headers

    def test_spoils_sheet_uses_sources(self, exporter, sample_tickets):
        """Test Spoils sheet uses SOURCE locations (not destinations)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Spoils"]
            
            # Check headers have spoils SOURCE locations
            headers = [cell.value for cell in ws[1]]
            assert "BECK SPOILS" in headers
            assert "NTX Spoils" in headers
            assert "UTX Spoils" in headers
            assert "MVP-TC1" in headers
            assert "MVP-TC2" in headers

    def test_import_sheet_content(self, exporter, sample_tickets):
        """Test Import sheet has correct material types."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Import"]
            
            # Check headers include import material types
            headers = [cell.value for cell in ws[1]]
            assert "3X5" in headers
            assert "ASPHALT" in headers
            assert "FLEXBASE" in headers
            assert "ROCK" in headers
            assert "UTILITY STONE" in headers
            assert "Grand Total" in headers

    def test_job_week_calculation(self, exporter):
        """Test job week calculation is correct."""
        test_date = date(2024, 10, 17)
        result = exporter._calculate_job_week(test_date)
        
        assert "Week" in result
        assert "End" in result
        assert "10/" in result  # October date

    def test_job_month_calculation(self, exporter):
        """Test job month calculation is correct."""
        test_date = date(2024, 10, 17)
        result = exporter._calculate_job_month(test_date)
        
        assert "October" in result
        assert "24" in result  # Year
        # Should be month 4 (July=1, Aug=2, Sep=3, Oct=4)
        assert result.startswith("004")

    def test_group_by_date(self, exporter, sample_tickets):
        """Test grouping tickets by date."""
        grouped = exporter._group_by_date(sample_tickets)
        
        assert "2024-10-17" in grouped
        assert "2024-10-18" in grouped
        assert "2024-10-19" in grouped
        
        # Oct 17 should have 3 tickets
        assert len(grouped["2024-10-17"]) == 3
        # Oct 18 should have 2 tickets
        assert len(grouped["2024-10-18"]) == 2
        # Oct 19 should have 2 tickets
        assert len(grouped["2024-10-19"]) == 2

    def test_header_styling(self, exporter, sample_tickets):
        """Test that headers are styled correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            exporter.export(sample_tickets, output_path)
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb["All Daily"]
            
            # Check first row has styling
            first_cell = ws.cell(1, 1)
            assert first_cell.font.bold is True
            assert first_cell.fill.start_color.rgb is not None

    def test_empty_tickets_list(self, exporter):
        """Test exporter handles empty ticket list."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_tracking.xlsx"
            
            # Should not raise exception
            exporter.export([], output_path)
            
            assert output_path.exists()
            
            wb = openpyxl.load_workbook(output_path)
            # Should still have 5 sheets with headers
            assert len(wb.sheetnames) == 5

    def test_custom_job_start_date(self):
        """Test exporter with custom job start date."""
        custom_date = date(2024, 1, 1)
        exporter = ExcelTrackingExporter(job_start_date=custom_date)
        
        assert exporter.job_start_date == custom_date
        
        # Test date calculation with custom start
        test_date = date(2024, 3, 15)
        result = exporter._calculate_job_month(test_date)
        
        # March should be month 3 (Jan=1, Feb=2, Mar=3)
        assert result.startswith("003")

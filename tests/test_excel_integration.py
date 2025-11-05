"""Integration tests for Excel exporter with database (Issue #12)."""
import tempfile
from datetime import date
from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.truck_tickets.database import create_all_tables, seed_reference_data
from src.truck_tickets.database.ticket_repository import TicketRepository
from src.truck_tickets.exporters.excel_exporter import ExcelTrackingExporter


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestExcelIntegration:
    """Integration tests for Excel exporter with database."""

    @pytest.fixture
    def session(self):
        """Create in-memory database with seeded data."""
        engine = create_engine("sqlite:///:memory:", echo=False)
        create_all_tables(engine)
        session_factory = sessionmaker(bind=engine)
        sess = session_factory()
        seed_reference_data(sess, job_code="24-105", job_name="PHMS New Pediatric Campus")
        try:
            yield sess
        finally:
            sess.close()
            engine.dispose()

    @pytest.fixture
    def repository(self, session):
        """Create repository instance."""
        return TicketRepository(session)

    def test_export_from_database_tickets(self, session, repository):
        """Test exporting tickets directly from database."""
        # Create some test tickets in database
        ticket_data = [
            {
                "ticket_number": "WM-40000001",
                "ticket_date": date(2024, 10, 17),
                "job_name": "24-105",
                "vendor_name": "WASTE_MANAGEMENT_LEWISVILLE",
                "material_name": "CLASS_2_CONTAMINATED",
                "source_name": "PODIUM",
                "destination_name": "WASTE_MANAGEMENT_LEWISVILLE",
                "ticket_type_name": "EXPORT",
                "quantity": 15.5,
                "quantity_unit": "TONS",
                "manifest_number": "WM-MAN-2024-001234",
                "file_id": "test.pdf",
                "file_page": 1,
                "extraction_confidence": 0.95
            },
            {
                "ticket_number": "WM-40000002",
                "ticket_date": date(2024, 10, 17),
                "job_name": "24-105",
                "vendor_name": "WASTE_MANAGEMENT_LEWISVILLE",
                "material_name": "NON_CONTAMINATED",
                "source_name": "SPG",
                "destination_name": "LDI_YARD",
                "ticket_type_name": "EXPORT",
                "quantity": 18.0,
                "quantity_unit": "TONS",
                "file_id": "test.pdf",
                "file_page": 2,
                "extraction_confidence": 0.92
            },
            {
                "ticket_number": "WM-40000003",
                "ticket_date": date(2024, 10, 18),
                "job_name": "24-105",
                "vendor_name": "WASTE_MANAGEMENT_LEWISVILLE",
                "material_name": "SPOILS",
                "source_name": "BECK_SPOILS",
                "destination_name": "WASTE_MANAGEMENT_LEWISVILLE",
                "ticket_type_name": "EXPORT",
                "quantity": 14.0,
                "quantity_unit": "TONS",
                "manifest_number": "WM-MAN-2024-555555",  # SPOILS requires manifest
                "file_id": "test.pdf",
                "file_page": 3,
                "extraction_confidence": 0.88
            },
        ]

        # Insert tickets
        for data in ticket_data:
            repository.create(**data)

        # Query tickets back from database
        from src.truck_tickets.models.sql_truck_ticket import TruckTicket
        db_tickets = session.query(TruckTicket).all()

        # Convert to dict format for exporter
        ticket_dicts = []
        for ticket in db_tickets:
            ticket_dicts.append({
                "ticket_date": ticket.ticket_date.strftime("%Y-%m-%d"),
                "ticket_number": ticket.ticket_number,
                "material": ticket.material.material_name if ticket.material else "UNKNOWN",
                "source": ticket.source.source_name if ticket.source else "UNKNOWN",
                "destination": ticket.destination.destination_name if ticket.destination else "UNKNOWN",
                "ticket_type": ticket.ticket_type.type_name if ticket.ticket_type else "UNKNOWN",
                "quantity": float(ticket.quantity) if ticket.quantity else 0.0,
                "vendor": ticket.vendor.vendor_name if ticket.vendor else "UNKNOWN"
            })

        # Export to Excel
        exporter = ExcelTrackingExporter(job_start_date=date(2024, 7, 1))
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "integration_test.xlsx"
            exporter.export(ticket_dicts, output_path)
            
            # Verify file was created
            assert output_path.exists()
            
            # Load and verify content
            wb = openpyxl.load_workbook(output_path)
            
            # Check All Daily sheet
            ws_all = wb["All Daily"]
            assert ws_all.max_row >= 2  # Header + at least 1 data row
            
            # Check Class2_Daily sheet
            ws_class2 = wb["Class2_Daily"]
            assert ws_class2.max_row >= 1  # At least header
            
            # Check Spoils sheet
            ws_spoils = wb["Spoils"]
            assert ws_spoils.max_row >= 1  # At least header

    def test_export_with_job_metrics(self, session, repository):
        """Test that job week/month calculations are correct in export."""
        # Create ticket
        repository.create(
            ticket_number="WM-50000001",
            ticket_date=date(2024, 10, 17),
            job_name="24-105",
            vendor_name="WASTE_MANAGEMENT_LEWISVILLE",
            material_name="CLASS_2_CONTAMINATED",
            source_name="PODIUM",
            destination_name="WASTE_MANAGEMENT_LEWISVILLE",
            ticket_type_name="EXPORT",
            quantity=15.5,
            quantity_unit="TONS",
            manifest_number="WM-MAN-2024-999999",
            file_id="test.pdf",
            file_page=1,
            extraction_confidence=0.95
        )

        # Query and convert
        from src.truck_tickets.models.sql_truck_ticket import TruckTicket
        ticket = session.query(TruckTicket).first()
        
        ticket_dict = {
            "ticket_date": ticket.ticket_date.strftime("%Y-%m-%d"),
            "ticket_number": ticket.ticket_number,
            "material": ticket.material.material_name if ticket.material else "UNKNOWN",
            "source": ticket.source.source_name if ticket.source else "UNKNOWN",
            "destination": ticket.destination.destination_name if ticket.destination else "UNKNOWN",
            "ticket_type": ticket.ticket_type.type_name if ticket.ticket_type else "UNKNOWN",
            "quantity": float(ticket.quantity) if ticket.quantity else 0.0,
            "vendor": ticket.vendor.vendor_name if ticket.vendor else "UNKNOWN"
        }

        # Export
        exporter = ExcelTrackingExporter(job_start_date=date(2024, 7, 1))
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "metrics_test.xlsx"
            exporter.export([ticket_dict], output_path)
            
            # Load and check job metrics
            wb = openpyxl.load_workbook(output_path)
            ws = wb["All Daily"]
            
            # Row 2 should have the data (row 1 is header)
            date_cell = ws.cell(2, 1).value
            day_cell = ws.cell(2, 2).value
            job_week_cell = ws.cell(2, 3).value
            job_month_cell = ws.cell(2, 4).value
            
            assert date_cell == "2024-10-17"
            assert day_cell == "Thu"
            assert "Week" in job_week_cell
            assert "October" in job_month_cell
            assert "004" in job_month_cell  # 4th month (July=1, Aug=2, Sep=3, Oct=4)

    def test_empty_database_export(self, session):
        """Test exporting when database has no tickets."""
        exporter = ExcelTrackingExporter(job_start_date=date(2024, 7, 1))
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "empty_test.xlsx"
            exporter.export([], output_path)
            
            # Should still create file with headers
            assert output_path.exists()
            
            wb = openpyxl.load_workbook(output_path)
            assert len(wb.sheetnames) == 5
            
            # Each sheet should have headers but no data
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                assert ws.max_row == 1  # Only header row

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / 'src'))
from doctr_process.ocr import reporting_utils


def test_export_issue_logs(tmp_path):
    cfg = {'output_dir': str(tmp_path), 'ticket_issues': True, 'issues_log': True}
    ticket_issues = [{'page': 1, 'issue': 'missing ticket'}]
    issues_log = [{'page': 1, 'issue_type': 'missing_field', 'field': 'ticket_number'}]
    reporting_utils.export_issue_logs(ticket_issues, issues_log, cfg)
    assert (tmp_path / 'logs' / 'ticket_issues.csv').exists()
    ti = pd.read_csv(tmp_path / 'logs' / 'ticket_issues.csv')
    assert ti.shape[0] == 1


def test_export_process_analysis(tmp_path):
    cfg = {'output_dir': str(tmp_path), 'process_analysis': True}
    records = [{'page': 1, 'ocr_time': 0.1, 'orientation_time': 0.2, 'orientation': 0}]
    reporting_utils.export_process_analysis(records, cfg)
    path = tmp_path / 'logs' / 'process_analysis.csv'
    assert path.exists()
    df = pd.read_csv(path)
    assert 'ocr_time' in df.columns


def test_condensed_ticket_report(tmp_path):
    cfg = {'output_dir': str(tmp_path), 'ticket_numbers_condensed_csv': True}
    rows = [
        {
            'file': '24-105_2025-07-30_Class2_Podium_WM.pdf',
            'page': 1,
            'vendor': 'ACME',
            'ticket_number': '123',
            'manifest_number': '14123456',
            'truck_number': 'TR1',
            'exception_reason': None,
            'image_path': 'img.png',
            'roi_image_path': 'roi.png'
        },
        # Include an invalid row to ensure colour-coding works
        {
            'file': '24-105_2025-07-30_Class2_Podium_WM.pdf',
            'page': 2,
            'vendor': 'ACME',
            'ticket_number': 'BAD',
            'manifest_number': '123',
            'truck_number': 'TR2',
            'exception_reason': None,
            'image_path': 'img2.png',
            'roi_image_path': 'roi2.png'
        },
        # Row with missing ticket and manifest numbers
        {
            'file': '24-105_2025-07-30_Class2_Podium_WM.pdf',
            'page': 3,
            'vendor': 'ACME',
            'ticket_number': '',
            'manifest_number': '',
            'truck_number': 'TR3',
            'exception_reason': None,
            'image_path': 'img3.png',
            'roi_image_path': 'roi3.png'
        }
    ]
    reporting_utils.create_reports(rows, cfg)
    path = tmp_path / 'logs' / 'ticket_number' / 'condensed_ticket_numbers.csv'
    assert path.exists()
    df = pd.read_csv(path)
    assert list(df.columns) == [
        'JobID',
        'Service Date',
        'Material',
        'Source',
        'Destination',
        'page',
        'vendor',
        'ticket_number',
        'ticket_number_valid',
        'manifest_number',
        'manifest_number_valid',
        'truck_number',
        'exception_reason',
        'image_path',
        'roi_image_path'
    ]
    assert df.iloc[0]['JobID'] == '24-105'

    # Validate Excel specific formatting
    excel_path = tmp_path / 'logs' / 'ticket_number' / 'condensed_ticket_numbers.xlsx'
    assert excel_path.exists()
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Image column should show filename with hyperlink
    img_cell = ws.cell(row=2, column=14)
    assert img_cell.value == 'img.png'
    assert img_cell.hyperlink.target == 'img.png'

    # Invalid ticket numbers should be highlighted and keep hyperlink
    invalid_ticket = ws.cell(row=3, column=8)
    assert invalid_ticket.hyperlink.target == 'roi2.png'
    assert invalid_ticket.fill.start_color.rgb.endswith('FFC7CE')

    # Invalid manifest numbers should also be highlighted and linked
    invalid_manifest = ws.cell(row=3, column=10)
    assert invalid_manifest.hyperlink.target == 'roi2.png'
    assert invalid_manifest.fill.start_color.rgb.endswith('BDD7EE')

    # Missing ticket numbers should show placeholder text with hyperlink
    missing_ticket = ws.cell(row=4, column=8)
    assert missing_ticket.value == 'missing'
    assert missing_ticket.hyperlink.target == 'roi3.png'
    assert missing_ticket.fill.start_color.rgb.endswith('FFEB9C')

    # Missing manifest numbers should show placeholder text with hyperlink
    missing_manifest = ws.cell(row=4, column=10)
    assert missing_manifest.value == 'missing'
    assert missing_manifest.hyperlink.target == 'roi3.png'
    assert missing_manifest.fill.start_color.rgb.endswith('F4B084')


def test_write_management_report(tmp_path):
    summary = {
        'JobID': '24-105',
        'Service Date': '2025-07-30',
        'Material': 'Class2',
        'Source': 'Podium',
        'Destination': 'WM',
        'total_pages': 2,
        'tickets_valid': 1,
        'tickets_invalid': 1,
        'tickets_missing': 0,
        'manifest_valid': 1,
        'manifest_review': 0,
        'manifest_invalid': 1,
    }
    vendors = [
        {
            'vendor': 'ACME',
            'page_count': 2,
            'vendor_doc_path': str(tmp_path / 'acme.pdf'),
        }
    ]
    cfg = reporting_utils.REPORTING_CFG.copy()
    cfg.update({
        'output_dir': str(tmp_path),
        'mgmt_report_pdf': False,
        'branding_logo_path': str(tmp_path / 'logo.png'),
    })
    reporting_utils.write_management_report(summary, vendors, cfg)
    report = tmp_path / 'management_report.xlsx'
    assert report.exists()
    from openpyxl import load_workbook
    wb = load_workbook(report)
    ws = wb.active
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == 'Vendor' and ws.cell(r, 2).value == 'Page Count':
            header_row = r
            break
    assert header_row is not None
    first_vendor = header_row + 1
    assert ws.cell(first_vendor, 1).value == 'ACME'
    doc_cell = ws.cell(first_vendor, 3)
    assert doc_cell.value == 'acme.pdf'
    assert doc_cell.hyperlink.target == str(tmp_path / 'acme.pdf')

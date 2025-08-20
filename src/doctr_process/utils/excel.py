# src/doctr_process/utils/excel.py
from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _cell_text(value: Any) -> str:
    """Convert a cell value to display text."""
    return "" if value is None else str(value)


def best_fit_column_widths(ws: Worksheet, max_width: int = 60, padding: int = 2) -> None:
    """
    Compute and apply best-fit widths for all columns in a worksheet.

    Args:
        ws: OpenPyXL worksheet to process.
        max_width: Maximum width in Excel column units (approx chars).
        padding: Extra width to add beyond the longest content.
    """
    # Track the longest text length per column index (1-based)
    longest: dict[int, int] = {}

    # Include header and data rows
    for row in ws.iter_rows(values_only=True):
        for col_idx, value in enumerate(row, start=1):
            text = _cell_text(value)
            # Handle multi-line cells by taking the longest line
            length = max((len(line) for line in text.splitlines()), default=0)
            if length > longest.get(col_idx, 0):
                longest[col_idx] = length

    # Apply widths (Excel "width" ~= number of characters)
    for col_idx, length in longest.items():
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max_width, length + padding)


def autosize_workbook(path: str | Path, max_width: int = 60, padding: int = 2) -> Path:
    """
    Auto-size columns for all sheets in an existing Excel workbook.

    Args:
        path: Path to an existing `.xlsx` file.
        max_width: Maximum width in Excel column units.
        padding: Extra width to add beyond the longest content.

    Returns:
        The saved workbook path.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the path is not an `.xlsx` file.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {p}")
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"Only `.xlsx` files are supported: {p}")

    wb = load_workbook(p)
    for ws in wb.worksheets:
        best_fit_column_widths(ws, max_width=max_width, padding=padding)
    wb.save(p)
    logger.info(f"Auto-sized Excel columns: {p}")
    return p


# Optional convenience for pandas usage.
def to_excel_autosize(
    df, path: str | Path, sheet_name: str = "Sheet1", index: bool = False, max_width: int = 60, padding: int = 2
) -> Path:
    """
    Write a pandas DataFrame to Excel and auto-size columns in the same pass.

    Args:
        df: pandas DataFrame.
        path: Output `.xlsx` path.
        sheet_name: Target sheet name.
        index: Whether to write the DataFrame index.
        max_width: Maximum width in Excel column units.
        padding: Extra width to add beyond the longest content.

    Returns:
        The saved workbook path.
    """
    import pandas as pd  # Local import to avoid hard dependency at module import time

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(p, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)
        ws = writer.sheets[sheet_name]
        best_fit_column_widths(ws, max_width=max_width, padding=padding)
    logger.info(f"Saved Excel with auto-sized columns: {p}")
    return p


# tests/test_excel_autosize.py
import pandas as pd
from openpyxl import load_workbook

from doctr_process.utils.excel import autosize_workbook


def test_autosize_workbook_sets_wider_columns(tmp_path):
    # Arrange
    df = pd.DataFrame(
        {
            "Short": ["a", "bb", "ccc"],
            "VeryLongHeaderNameForColumn": ["x" * 10, "y" * 25, "z" * 40],
        }
    )
    out = tmp_path / "out.xlsx"
    df.to_excel(out, index=False)

    # Act
    autosize_workbook(out)

    # Assert
    wb = load_workbook(out)
    ws = wb.active
    short_w = ws.column_dimensions["A"].width or 0
    long_w = ws.column_dimensions["B"].width or 0
    # Long column should be wider than short; also ensure it's reasonably wide
    assert long_w > short_w
    assert long_w >= 20
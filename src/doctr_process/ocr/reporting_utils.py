import csv
from html import escape
import html
import logging
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd

def _ensure_hyperlink_style(cell, link_path: str, display_text: str = None) -> None:
    """Ensure consistent hyperlink styling across all Excel reports."""
    try:
        from openpyxl.styles import Font
        
        if link_path:
            cell.hyperlink = link_path
            # Use display text if provided, otherwise use the filename
            if display_text:
                cell.value = display_text
            else:
                cell.value = Path(link_path).name if Path(link_path).is_absolute() else link_path
            
            # Ensure hyperlink styling is applied
            cell.font = Font(color="0563C1", underline="single")
            # Alternative: cell.style = "Hyperlink" - but this may override other formatting
    except Exception:
        # Fallback to just setting the value without hyperlink
        cell.value = display_text or (Path(link_path).name if link_path else "")


# Standard column order for consistent business reports
STANDARD_COLUMN_ORDER = [
    "JobID",
    "Service Date", 
    "Material",
    "Source",
    "Destination",
    "page",
    "vendor",
    "ticket_number",
    "ticket_number_valid",
    "manifest_number", 
    "manifest_number_valid",
    "truck_number",
    "exception_reason",
    "image_path",
    "roi_image_path",
    "manifest_roi_image_path",
]

REPORTING_CFG = {
    "branding_company_name": "Lindamood Demolition, Inc.",
    "branding_logo_path": str(
        Path(__file__).parent / "assets" / "branding" / "lindamood_logo.png"
    ),
    "report_author": "B. Atkins",
    "script_version": "1.2.0",
    "mgmt_report_xlsx": True,
    "mgmt_report_pdf": True,
    "pdf_export": {"method": "auto"},
}


def _parse_log_line(line: str) -> List[str]:
    """Parse a log line produced by doctr_ocr_to_csv."""
    parts = line.strip().split(",", 4)
    if len(parts) < 5:
        # Fallback if message contains commas or format unexpected
        dt = parts[0] if parts else ""
        level = parts[1] if len(parts) > 1 else ""
        file = parts[2] if len(parts) > 2 else ""
        lineno = parts[3] if len(parts) > 3 else ""
        # Fallback: assign message if present, else use the whole line
        msg = parts[4] if len(parts) > 4 else line.strip()
        return [dt, level, file, lineno, msg]
    return parts[:5]


def export_logs_to_csv(log_file_path: str, output_csv_path: str) -> None:
    """Convert ``error.log`` to a structured CSV."""
    rows = []
    try:
        with open(log_file_path, encoding="utf-8") as f:
            for line in f:
                dt, level, file, lineno, msg = _parse_log_line(line)
                rows.append(
                    {
                        "datetime": dt,
                        "level": level,
                        "file": file,
                        "line": lineno,
                        "message": msg,
                    }
                )
    except FileNotFoundError:
        return

    with open(output_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["datetime", "level", "file", "line", "message"],
        )
        writer.writeheader()
        writer.writerows(rows)


def export_logs_to_html(log_file_path: str, output_html_path: str) -> None:
    """Convert ``error.log`` to a simple HTML table."""
    try:
        with open(log_file_path, encoding="utf-8") as f:
            lines = f.readlines()
    except FileNotFoundError:
        lines = []

    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write("<html><body><table border='1'>")
        f.write(
            "<tr><th>DateTime</th><th>Level</th><th>File</th><th>Line</th><th>Message</th></tr>"
        )
        for line in lines:
            dt, level, file, lineno, msg = _parse_log_line(line)
            f.write("<tr>")
            f.write(f"<td>{html.escape(dt)}</td>")
            f.write(f"<td>{html.escape(level)}</td>")
            f.write(f"<td>{html.escape(file)}</td>")
            f.write(f"<td>{html.escape(lineno)}</td>")
            f.write(f"<td>{html.escape(msg)}</td>")
            f.write("</tr>")
        f.write("</table></body></html>")


def _report_path(cfg: Dict[str, Any], key: str, sub_path: str) -> str | None:
    """Resolve a report path from ``output_dir`` and ``key``."""
    val = cfg.get(key)
    if isinstance(val, bool) or val is None:
        if not val:
            return None
        base = Path(cfg.get("output_dir", "./outputs")) / "logs" / sub_path
        return str(base)
    return str(val)


def generate_business_summary(cfg: Dict[str, Any]) -> None:
    """Generate a business-friendly summary of all pipeline outputs."""
    output_dir = Path(cfg.get("output_dir", "./outputs"))
    summary_path = output_dir / "business_summary.txt"
    
    try:
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("DocTR Process - Business Summary Report\n")
            f.write("=" * 45 + "\n\n")
            f.write(f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}\n")
            f.write(f"Output Directory: {output_dir}\n\n")
            
            # Management Report
            mgmt_report = output_dir / "management_report.xlsx"
            if mgmt_report.exists():
                f.write("📊 MANAGEMENT REPORT\n")
                f.write(f"   Excel: {mgmt_report.name}\n")
                if (output_dir / "management_report.pdf").exists():
                    f.write(f"   PDF: management_report.pdf\n")
                f.write("\n")
            
            # Data Exports
            f.write("📁 DATA EXPORTS\n")
            csv_files = list(output_dir.rglob("*.csv"))
            if csv_files:
                for csv_file in sorted(csv_files):
                    rel_path = csv_file.relative_to(output_dir)
                    f.write(f"   📄 {rel_path}\n")
            else:
                f.write("   No CSV files generated\n")
            f.write("\n")
            
            # Excel Reports
            f.write("📊 EXCEL REPORTS\n")
            xlsx_files = list(output_dir.rglob("*.xlsx"))
            if xlsx_files:
                for xlsx_file in sorted(xlsx_files):
                    rel_path = xlsx_file.relative_to(output_dir)
                    f.write(f"   📊 {rel_path}\n")
            else:
                f.write("   No Excel files generated\n")
            f.write("\n")
            
            # Vendor Documents
            f.write("📄 VENDOR DOCUMENTS\n")
            vendor_docs = list(output_dir.rglob("vendor_docs/*.pdf"))
            if vendor_docs:
                for doc in sorted(vendor_docs):
                    f.write(f"   📄 {doc.name}\n")
                
                # Check for combined PDF
                combined_pdfs = [d for d in vendor_docs if "combined" in d.name.lower()]
                if combined_pdfs:
                    f.write(f"   🔗 Combined: {combined_pdfs[0].name}\n")
            else:
                f.write("   No vendor documents generated\n")
            f.write("\n")
            
            # Usage Instructions
            f.write("📋 USAGE INSTRUCTIONS\n")
            f.write("   1. Review management_report.xlsx for high-level metrics\n")
            f.write("   2. Check vendor documents for individual PDFs\n")
            f.write("   3. Use CSV files for detailed data analysis\n")
            f.write("   4. Excel reports include hyperlinks to source images\n")
            f.write("\n")
            
            f.write("For technical details, see artifact_summary.json\n")
        
        logging.info("Generated business summary: %s", summary_path)
        
    except Exception as e:
        logging.error("Failed to generate business summary: %s", str(e))


def export_log_reports(cfg: Dict[str, Any]) -> None:
    """Export ``error.log`` to CSV/HTML if enabled in ``cfg``."""
    log_file = "error.log"
    csv_path = _report_path(cfg, "log_csv", "log_report.csv")
    if csv_path:
        Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
        export_logs_to_csv(log_file, csv_path)
    html_path = _report_path(cfg, "log_html", "log_report.html")
    if html_path:
        Path(html_path).parent.mkdir(parents=True, exist_ok=True)
        export_logs_to_html(log_file, html_path)


def get_manifest_validation_status(manifest_number: str | None) -> str:
    """Return validation status for a manifest number."""
    if not manifest_number:
        return "invalid"
    manifest_number = str(manifest_number)
    if re.fullmatch(r"14\d{6}", manifest_number):
        return "valid"
    if len(manifest_number) >= 7:
        return "review"
    return "invalid"


def get_ticket_validation_status(ticket_number: str | None) -> str:
    """Return validation status for a ticket number."""
    if not ticket_number:
        return "invalid"
    ticket_number = str(ticket_number)
    if re.fullmatch(r"\d+", ticket_number):
        return "valid"
    return "review"


def _parse_filename_metadata(file_path: str) -> Dict[str, str]:
    """Extract job related fields from an input ``file_path``.

    Filenames are expected to follow the pattern
    ``JobID_ServiceDate_Material_Source_Destination[_...]``.  A trailing
    ``_NNN`` segment (original page count) is stripped if present.  Missing
    segments yield empty strings.
    """

    stem = Path(file_path).stem
    stem = re.sub(r"_(\d+)$", "", stem)
    parts = stem.split("_")
    fields = ["JobID", "Service Date", "Material", "Source", "Destination"]
    meta = {k: "" for k in fields}
    for key, value in zip(fields, parts):
        meta[key] = value
    return meta


def _make_vendor_doc_path(
        file_path: str, vendor: str, page_count: int, cfg: Dict[str, Any]
) -> str:
    """Return the expected vendor document path for ``file_path`` and ``vendor``."""

    # Import formatting helpers lazily to avoid importing optional dependencies
    # such as PyPDF2 when this function isn't used.  This keeps modules that
    # merely import ``reporting_utils`` lightweight.
    from src.doctr_process.processor.filename_utils import (
        format_output_filename,
        format_output_filename_camel,
        format_output_filename_lower,
        format_output_filename_snake,
        format_output_filename_preserve,
        parse_input_filename_fuzzy,
        sanitize_vendor_name,
    )

    # Sanitize vendor to prevent path traversal
    safe_vendor = re.sub(r"[\\/]", "_", vendor)
    vendor_dir = Path(cfg.get("output_dir", "./outputs")) / "vendor_docs"
    fmt_style = cfg.get("file_format", "preserve").lower()
    format_func = {
        "camel": format_output_filename_camel,
        "caps": format_output_filename,
        "lower": format_output_filename_lower,
        "snake": format_output_filename_snake,
        "preserve": format_output_filename_preserve,
    }.get(fmt_style, format_output_filename_preserve)

    meta = parse_input_filename_fuzzy(file_path)
    vendor_clean = sanitize_vendor_name(vendor)
    fmt = cfg.get("vendor_doc_format", "pdf")
    out_name = format_func(vendor_clean, page_count, meta, fmt)
    return str(vendor_dir / out_name)


def create_reports(rows: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    """Write combined, deduped and summary CSV reports."""
    if not rows:
        return

    # Combined OCR dump
    out_path = _report_path(cfg, "output_csv", "ocr/combined_results.csv")
    if out_path:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        pd.DataFrame(rows).to_csv(str(out_path), index=False)

    df = pd.DataFrame(rows)
    df["ticket_valid"] = df["ticket_number"].apply(get_ticket_validation_status)
    df["manifest_valid"] = df["manifest_number"].apply(get_manifest_validation_status)

    # Page-level fields with validation
    page_fields_path = _report_path(cfg, "page_fields_csv", "ocr/page_fields.csv")
    if page_fields_path:
        os.makedirs(os.path.dirname(page_fields_path), exist_ok=True)
        df.to_csv(str(page_fields_path), index=False)

    # Mark duplicate tickets
    ticket_counts = df.groupby(["vendor", "ticket_number"]).size().rename("count")
    df = df.join(ticket_counts, on=["vendor", "ticket_number"])
    df["duplicate_ticket"] = df["count"] > 1

    ticket_numbers_path = _report_path(
        cfg, "ticket_numbers_csv", "ocr/combined_ticket_numbers.csv"
    )
    if ticket_numbers_path:
        os.makedirs(os.path.dirname(ticket_numbers_path), exist_ok=True)
        df.drop(columns=["count"]).to_csv(str(ticket_numbers_path), index=False)

    condensed_path = _report_path(
        cfg,
        "ticket_numbers_condensed_csv",
        "ticket_number/condensed_ticket_numbers.csv",
    )
    if condensed_path:
        os.makedirs(os.path.dirname(condensed_path), exist_ok=True)
        condensed_records: List[Dict[str, Any]] = []
        for _, row in df.iterrows():
            meta = _parse_filename_metadata(row.get("file", ""))
            record = {
                **meta,
                "page": row.get("page"),
                "vendor": row.get("vendor"),
                "ticket_number": row.get("ticket_number"),
                "ticket_number_valid": row.get("ticket_valid"),
                "manifest_number": row.get("manifest_number"),
                "manifest_number_valid": row.get("manifest_valid"),
                "truck_number": row.get("truck_number"),
                "exception_reason": row.get("exception_reason"),
                "image_path": row.get("image_path"),
                "roi_image_path": row.get("roi_image_path"),
                "manifest_roi_image_path": row.get("manifest_roi_image_path"),
            }
            condensed_records.append(record)
        columns = [col for col in STANDARD_COLUMN_ORDER if col in condensed_df.columns]
        # Ensure all required columns exist, add missing ones with None values
        for col in STANDARD_COLUMN_ORDER:
            if col not in condensed_df.columns:
                condensed_df[col] = None
        condensed_df = pd.DataFrame(condensed_records)
        condensed_df[columns].to_csv(str(condensed_path), index=False)
        excel_path = Path(condensed_path).with_suffix(".xlsx")
        condensed_df[columns].to_excel(str(excel_path), index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            wb = load_workbook(str(excel_path))
            ws = wb.active
            invalid_ticket_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            missing_ticket_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
            invalid_manifest_fill = PatternFill(
                start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
            )
            missing_manifest_fill = PatternFill(
                start_color="F4B084", end_color="F4B084", fill_type="solid"
            )
            t_col = columns.index("ticket_number") + 1
            m_col = columns.index("manifest_number") + 1
            img_col = columns.index("image_path") + 1
            roi_col = columns.index("roi_image_path") + 1

            def _set_cell(
                    row: int, col: int, value: Any, link: str | None = None, fill=None
            ) -> None:
                """Populate ``ws`` cell ensuring fill is applied last."""
                c = ws.cell(row=row, column=col)
                c.value = value
                if link:
                    _ensure_hyperlink_style(c, link, str(value))
                if fill is not None:
                    c.fill = fill

            for idx, rec in condensed_df.iterrows():
                r = idx + 2
                # Replace image path columns with filename hyperlinks
                img = rec.get("image_path")
                if img:
                    fname = Path(img).name
                    _set_cell(r, img_col, fname, img)
                roi = rec.get("roi_image_path")
                if roi:
                    fname = Path(roi).name
                    _set_cell(r, roi_col, fname, roi)

                # Highlight missing/invalid ticket numbers with distinct colours
                ticket = rec.get("ticket_number")
                if rec.get("ticket_number_valid") != "valid":
                    link = roi if roi else None
                    value = ticket if ticket else "missing"
                    fill = invalid_ticket_fill if ticket else missing_ticket_fill
                    _set_cell(r, t_col, value, link, fill)

                # Highlight missing/invalid manifest numbers with distinct colours
                manifest = rec.get("manifest_number")
                if rec.get("manifest_number_valid") != "valid":
                    m_link = rec.get("manifest_roi_image_path") or roi
                    value = manifest if manifest else "missing"
                    fill = invalid_manifest_fill if manifest else missing_manifest_fill
                    _set_cell(r, m_col, value, m_link, fill)

            # Prevent path traversal by ensuring excel_path is within output_dir
            output_dir = Path(cfg.get("output_dir", "./outputs"))
            excel_path = output_dir / Path(excel_path).name
            wb.save(str(excel_path))
        except ImportError:
            # openpyxl is an optional dependency; if it's missing we simply
            # emit the basic Excel file without hyperlink/colour enhancements.
            pass

    # Ticket/manifest exception logs
    ticket_exc = df[df["ticket_number"].isna() | (df["ticket_number"] == "")]
    ticket_exc_path = _report_path(
        cfg,
        "ticket_number_exceptions_csv",
        "ticket_number/ticket_number_exceptions.csv",
    )
    if ticket_exc_path:
        os.makedirs(os.path.dirname(ticket_exc_path), exist_ok=True)
        ticket_exc.to_csv(str(ticket_exc_path), index=False)

    manifest_exc = df[df["manifest_valid"] != "valid"]

    manifest_exc_path = _report_path(
        cfg,
        "manifest_number_exceptions_csv",
        "manifest_number/manifest_number_exceptions.csv",
    )

    if manifest_exc_path:
        os.makedirs(os.path.dirname(manifest_exc_path), exist_ok=True)
        manifest_exc.to_csv(str(manifest_exc_path), index=False)

    # Summary report
    summary_path = _report_path(cfg, "summary_report", "summary/summary.csv")
    if summary_path:
        os.makedirs(os.path.dirname(summary_path), exist_ok=True)
        ticket_valid_count = int((df["ticket_valid"] == "valid").sum())
        ticket_missing_count = int(ticket_exc.shape[0])
        ticket_invalid_count = len(df) - ticket_valid_count - ticket_missing_count
        summary_csv = {
            "total_pages": len(df),
            "tickets_missing": ticket_missing_count,
            "manifest_valid": int((df["manifest_valid"] == "valid").sum()),
            "manifest_review": int((df["manifest_valid"] == "review").sum()),
            "manifest_invalid": int((df["manifest_valid"] == "invalid").sum()),
        }
        summary_df = pd.DataFrame([summary_csv])

        vendor_counts = (
            df.groupby(["file", "vendor"]).size().reset_index(name="page_count")
        )
        if not vendor_counts.empty:
            vendor_counts["vendor_doc_path"] = vendor_counts.apply(
                lambda r: _make_vendor_doc_path(
                    r["file"], r["vendor"], r["page_count"], cfg
                ),
                axis=1,
            )

        with open(summary_path, "w", newline="", encoding="utf-8") as f:
            summary_df.to_csv(f, index=False)
            if not vendor_counts.empty:
                f.write("\n")
                vendor_counts.to_csv(f, index=False)

        summary_meta = {}
        if not df.empty:
            summary_meta = _parse_filename_metadata(df.iloc[0]["file"])
        summary_data = {
            **summary_meta,
            **summary_csv,
            "tickets_valid": ticket_valid_count,
            "tickets_invalid": ticket_invalid_count,
        }
        vendor_records = []
        if not vendor_counts.empty:
            vendor_records = vendor_counts[
                ["vendor", "page_count", "vendor_doc_path"]
            ].to_dict("records")
        report_cfg = REPORTING_CFG.copy()
        report_cfg["output_dir"] = Path(summary_path).parent
        try:
            write_management_report(summary_data, vendor_records, report_cfg)
        except Exception:
            logging.exception("Failed to write management report")


def _calculate_processing_stats(summary: Dict[str, Any], vendors: List[Dict[str, Any]]) -> List[tuple[str, Any]]:
    """Calculate processing statistics for management report."""
    total_pages = summary.get("total_pages", 0)
    valid_tickets = summary.get("tickets_valid", 0)
    valid_manifests = summary.get("manifest_valid", 0)
    vendor_count = len(vendors)
    
    # Calculate accuracy rates
    ticket_accuracy = (valid_tickets / total_pages * 100) if total_pages > 0 else 0
    manifest_accuracy = (valid_manifests / total_pages * 100) if total_pages > 0 else 0
    
    # Calculate pages per vendor
    avg_pages_per_vendor = (total_pages / vendor_count) if vendor_count > 0 else 0
    
    return [
        ("Vendor Count", vendor_count),
        ("Avg Pages per Vendor", f"{avg_pages_per_vendor:.1f}"),
        ("Ticket Accuracy", f"{ticket_accuracy:.1f}%"),
        ("Manifest Accuracy", f"{manifest_accuracy:.1f}%"),
    ]


def write_management_report(
        summary: Dict[str, Any], vendors: List[Dict[str, Any]], cfg: Dict[str, Any]
) -> None:
    """Create a formatted Excel management report and optional PDF."""
    if not cfg.get("mgmt_report_xlsx"):
        return

    output_dir = Path(cfg.get("output_dir", "./outputs"))
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "management_report.xlsx"

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    try:  # pragma: no cover - image support is optional
        from openpyxl.drawing.image import Image as XLImage
    except Exception:  # pragma: no cover - optional dependency
        XLImage = None

    wb = Workbook()
    ws = wb.active
    row = 1

    logo_path = cfg.get("branding_logo_path")
    if XLImage and logo_path and Path(logo_path).is_file():
        try:
            img = XLImage(logo_path)
            max_w = 600
            if img.width and img.width > max_w:
                ratio = max_w / img.width
                img.width = max_w
                img.height = img.height * ratio
            ws.add_image(img, "A1")
            row = int((img.height or 0) / 20) + 2
        except Exception:
            pass

    company = cfg.get("branding_company_name")
    if company:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=company)
        c.font = Font(bold=True, size=14)
        c.alignment = Alignment(horizontal="center")
        row += 1
    
    # Add report title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    title_cell = ws.cell(row=row, column=1, value="OCR Processing Management Report")
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    row += 2

    header_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_two_col_section(title: str, items: List[tuple[str, Any]]) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Field").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
        for col in (1, 2):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.border = border
        row += 1
        for field, value in items:
            ws.cell(row=row, column=1, value=field).border = border
            ws.cell(row=row, column=2, value=value).border = border
            row += 1
        row += 1

    section1 = [
        ("Job", summary.get("JobID") or summary.get("Job")),
        ("Date", summary.get("Service Date") or summary.get("Date")),
        ("Material", summary.get("Material")),
        ("Source", summary.get("Source")),
        ("Destination", summary.get("Destination")),
    ]
    write_two_col_section("Summary", section1)

    section2 = [
        ("Total Pages", summary.get("total_pages")),
        ("Valid Tickets", summary.get("tickets_valid")),
        ("Invalid Tickets", summary.get("tickets_invalid")),
        ("Missing Tickets", summary.get("tickets_missing")),
    ]
    write_two_col_section("Ticket Summary", section2)

    section3 = [
        ("Manifest Valid", summary.get("manifest_valid")),
        ("Manifest Review", summary.get("manifest_review")),
        ("Manifest Invalid", summary.get("manifest_invalid")),
    ]
    write_two_col_section("Manifest Summary", section3)

    # Add processing statistics section
    processing_stats = _calculate_processing_stats(summary, vendors)
    write_two_col_section("Processing Statistics", processing_stats)

    ws.cell(row=row, column=1, value="Vendor Details").font = Font(bold=True, size=12)
    row += 1
    headers = ["Vendor", "Page Count", "Document"]
    header_row = row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
    row += 1
    for rec in vendors:
        ws.cell(row=row, column=1, value=rec.get("vendor")).border = border
        ws.cell(row=row, column=2, value=rec.get("page_count")).border = border
        doc_path = rec.get("vendor_doc_path")
        filename = Path(doc_path).name if doc_path else ""
        cell = ws.cell(row=row, column=3, value=filename)
        if doc_path:
            _ensure_hyperlink_style(cell, doc_path, filename)
        cell.border = border
        row += 1
    ws.freeze_panes = f"A{header_row + 1}"

    for col in range(1, 4):
        max_len = 0
        for r in range(1, row + 1):
            val = ws.cell(r, col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    footer = f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S} • By: {cfg.get('report_author')} • Version: {cfg.get('script_version')}"
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=3)
    fcell = ws.cell(row=row + 1, column=1, value=footer)
    fcell.alignment = Alignment(horizontal="center")

    wb.save(str(xlsx_path))

    def _export_pdf(xlsx: Path) -> None:
        method = cfg.get("pdf_export", {}).get("method", "auto")
        if method in ("auto", "excel") and sys.platform.startswith("win"):
            try:
                import win32com.client  # type: ignore

                excel = win32com.client.Dispatch("Excel.Application")
                wb = excel.Workbooks.Open(str(xlsx))
                wb.ExportAsFixedFormat(0, str(xlsx.with_suffix(".pdf")))
                wb.Close(False)
                excel.Quit()
                return
            except Exception:
                if method == "excel":
                    logging.warning("Excel PDF export failed")
        if method in ("auto", "libreoffice"):
            soffice = shutil.which("soffice")
            if soffice:
                try:
                    subprocess.run(
                        [
                            soffice,
                            "--headless",
                            "--convert-to",
                            "pdf",
                            str(xlsx),
                            "--outdir",
                            str(output_dir),
                        ],
                        check=True,
                    )
                    return
                except Exception:
                    if method == "libreoffice":
                        logging.warning("LibreOffice PDF export failed")
        logging.warning(
            "Management report PDF export skipped; required tools not available"
        )

    if cfg.get("mgmt_report_pdf"):
        _export_pdf(xlsx_path)


def export_preflight_exceptions(
        exceptions: List[Dict[str, Any]], cfg: Dict[str, Any]
) -> None:
    """Write preflight exception rows to CSV if enabled."""
    if not exceptions:
        return
    out_path = _report_path(
        cfg,
        "preflight_exceptions_csv",
        "preflight/preflight_exceptions.csv",
    )
    if out_path:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        pd.DataFrame(exceptions).to_csv(str(out_path), index=False)


def export_issue_logs(
        ticket_issues: List[Dict[str, Any]],
        issues_log: List[Dict[str, Any]],
        cfg: Dict[str, Any],
) -> None:
    """Write detailed issue logs if enabled."""
    ti_path = _report_path(cfg, "ticket_issues", "ticket_issues.csv")
    if ti_path and ticket_issues:
        os.makedirs(os.path.dirname(ti_path), exist_ok=True)
        pd.DataFrame(ticket_issues).to_csv(str(ti_path), index=False)

    il_path = _report_path(cfg, "issues_log", "issues_log.csv")
    if il_path and issues_log:
        os.makedirs(os.path.dirname(il_path), exist_ok=True)
        pd.DataFrame(issues_log).to_csv(str(il_path), index=False)


def export_process_analysis(records: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    """Write per-page OCR timing metrics."""
    path = _report_path(cfg, "process_analysis", "process_analysis.csv")
    if not path or not records:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    pd.DataFrame(records).to_csv(str(path), index=False)

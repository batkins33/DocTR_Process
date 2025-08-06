import csv
import html
import os
import re
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from doctr_process.processor.filename_utils import (
    format_output_filename,
    format_output_filename_camel,
    format_output_filename_lower,
    format_output_filename_snake,
    format_output_filename_preserve,
    parse_input_filename_fuzzy,
    sanitize_vendor_name,
)


def _parse_log_line(line: str) -> List[str]:
    """Parse a log line produced by doctr_ocr_to_csv."""
    parts = line.strip().split(',', 4)
    if len(parts) < 5:
        # Fallback if message contains commas or format unexpected
        dt = parts[0] if parts else ''
        level = parts[1] if len(parts) > 1 else ''
        file = parts[2] if len(parts) > 2 else ''
        lineno = parts[3] if len(parts) > 3 else ''
        msg = parts[4] if len(parts) > 4 else line.strip()
        return [dt, level, file, lineno, msg]
    return parts[:5]


def export_logs_to_csv(log_file_path: str, output_csv_path: str) -> None:
    """Convert ``error.log`` to a structured CSV."""
    rows = []
    try:
        with open(log_file_path, encoding='utf-8') as f:
            for line in f:
                dt, level, file, lineno, msg = _parse_log_line(line)
                rows.append(
                    {
                        "datetime": dt,
                        "level": level,
                        "file": file,
                        "line": lineno,
                        "message": msg,
                    }
                )
    except FileNotFoundError:
        return

    with open(output_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["datetime", "level", "file", "line", "message"],
        )
        writer.writeheader()
        writer.writerows(rows)


def export_logs_to_html(log_file_path: str, output_html_path: str) -> None:
    """Convert ``error.log`` to a simple HTML table."""
    try:
        with open(log_file_path, encoding='utf-8') as f:
            lines = f.readlines()
    except FileNotFoundError:
        lines = []

    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write("<html><body><table border='1'>")
        f.write(
            "<tr><th>DateTime</th><th>Level</th><th>File</th><th>Line</th><th>Message</th></tr>"
        )
        for line in lines:
            dt, level, file, lineno, msg = _parse_log_line(line)
            f.write("<tr>")
            f.write(f"<td>{html.escape(dt)}</td>")
            f.write(f"<td>{html.escape(level)}</td>")
            f.write(f"<td>{html.escape(file)}</td>")
            f.write(f"<td>{html.escape(lineno)}</td>")
            f.write(f"<td>{html.escape(msg)}</td>")
            f.write("</tr>")
        f.write("</table></body></html>")



def _report_path(cfg: Dict[str, Any], key: str, sub_path: str) -> str | None:
    """Resolve a report path from ``output_dir`` and ``key``."""
    val = cfg.get(key)
    if isinstance(val, bool) or val is None:
        if not val:
            return None
        base = Path(cfg.get("output_dir", "./outputs")) / "logs" / sub_path
        return str(base)
    return str(val)


def export_log_reports(cfg: Dict[str, Any]) -> None:
    """Export ``error.log`` to CSV/HTML if enabled in ``cfg``."""
    log_file = "error.log"
    csv_path = _report_path(cfg, "log_csv", "log_report.csv")
    if csv_path:
        Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
        export_logs_to_csv(log_file, csv_path)
    html_path = _report_path(cfg, "log_html", "log_report.html")
    if html_path:
        Path(html_path).parent.mkdir(parents=True, exist_ok=True)
        export_logs_to_html(log_file, html_path)


def get_manifest_validation_status(manifest_number: str | None) -> str:
    """Return validation status for a manifest number."""
    if not manifest_number:
        return "invalid"
    manifest_number = str(manifest_number)
    if re.fullmatch(r"14\d{6}", manifest_number):
        return "valid"
    if len(manifest_number) >= 7:
        return "review"
    return "invalid"


def get_ticket_validation_status(ticket_number: str | None) -> str:
    """Return validation status for a ticket number."""
    if not ticket_number:
        return "invalid"
    ticket_number = str(ticket_number)
    if re.fullmatch(r"\d+", ticket_number):
        return "valid"
    return "review"


def _parse_filename_metadata(file_path: str) -> Dict[str, str]:
    """Extract job related fields from an input ``file_path``.

    Filenames are expected to follow the pattern
    ``JobID_ServiceDate_Material_Source_Destination[_...]``.  A trailing
    ``_NNN`` segment (original page count) is stripped if present.  Missing
    segments yield empty strings.
    """

    stem = Path(file_path).stem
    stem = re.sub(r"_(\d+)$", "", stem)
    parts = stem.split("_")
    fields = ["JobID", "Service Date", "Material", "Source", "Destination"]
    meta = {k: "" for k in fields}
    for key, value in zip(fields, parts):
        meta[key] = value
    return meta


def _make_vendor_doc_path(
    file_path: str, vendor: str, page_count: int, cfg: Dict[str, Any]
) -> str:
    """Return the expected vendor document path for ``file_path`` and ``vendor``."""

    vendor_dir = Path(cfg.get("output_dir", "./outputs")) / "vendor_docs"
    fmt_style = cfg.get("file_format", "preserve").lower()
    format_func = {
        "camel": format_output_filename_camel,
        "caps": format_output_filename,
        "lower": format_output_filename_lower,
        "snake": format_output_filename_snake,
        "preserve": format_output_filename_preserve,
    }.get(fmt_style, format_output_filename_preserve)

    meta = parse_input_filename_fuzzy(file_path)
    vendor_clean = sanitize_vendor_name(vendor)
    fmt = cfg.get("vendor_doc_format", "pdf")
    out_name = format_func(vendor_clean, page_count, meta, fmt)
    return str(vendor_dir / out_name)


def create_reports(rows: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    """Write combined, deduped and summary CSV reports."""
    if not rows:
        return

    # Combined OCR dump
    out_path = _report_path(cfg, "output_csv", "ocr/combined_results.csv")
    if out_path:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        pd.DataFrame(rows).to_csv(out_path, index=False)

    df = pd.DataFrame(rows)
    df["ticket_valid"] = df["ticket_number"].apply(get_ticket_validation_status)
    df["manifest_valid"] = df["manifest_number"].apply(get_manifest_validation_status)

    # Page-level fields with validation
    page_fields_path = _report_path(cfg, "page_fields_csv", "ocr/page_fields.csv")
    if page_fields_path:
        os.makedirs(os.path.dirname(page_fields_path), exist_ok=True)
        df.to_csv(page_fields_path, index=False)

    # Mark duplicate tickets
    ticket_counts = df.groupby(["vendor", "ticket_number"]).size().rename("count")
    df = df.join(ticket_counts, on=["vendor", "ticket_number"])
    df["duplicate_ticket"] = df["count"] > 1

    ticket_numbers_path = _report_path(cfg, "ticket_numbers_csv", "ocr/combined_ticket_numbers.csv")
    if ticket_numbers_path:
        os.makedirs(os.path.dirname(ticket_numbers_path), exist_ok=True)
        df.drop(columns=["count"]).to_csv(ticket_numbers_path, index=False)

    condensed_path = _report_path(
        cfg,
        "ticket_numbers_condensed_csv",
        "ticket_number/condensed_ticket_numbers.csv",
    )
    if condensed_path:
        os.makedirs(os.path.dirname(condensed_path), exist_ok=True)
        condensed_records: List[Dict[str, Any]] = []
        for _, row in df.iterrows():
            meta = _parse_filename_metadata(row.get("file", ""))
            record = {
                **meta,
                "page": row.get("page"),
                "vendor": row.get("vendor"),
                "ticket_number": row.get("ticket_number"),
                "ticket_number_valid": row.get("ticket_valid"),
                "manifest_number": row.get("manifest_number"),
                "manifest_number_valid": row.get("manifest_valid"),
                "truck_number": row.get("truck_number"),
                "exception_reason": row.get("exception_reason"),
                "image_path": row.get("image_path"),
                "roi_image_path": row.get("roi_image_path"),
                "manifest_roi_image_path": row.get("manifest_roi_image_path"),
            }
            condensed_records.append(record)
        columns = [
            "JobID",
            "Service Date",
            "Material",
            "Source",
            "Destination",
            "page",
            "vendor",
            "ticket_number",
            "ticket_number_valid",
            "manifest_number",
            "manifest_number_valid",
            "truck_number",
            "exception_reason",
            "image_path",
            "roi_image_path",
        ]
        condensed_df = pd.DataFrame(condensed_records)
        condensed_df[columns].to_csv(condensed_path, index=False)
        excel_path = Path(condensed_path).with_suffix(".xlsx")
        condensed_df[columns].to_excel(excel_path, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            from openpyxl.worksheet.hyperlink import Hyperlink

            wb = load_workbook(excel_path)
            ws = wb.active
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            t_col = columns.index("ticket_number") + 1
            m_col = columns.index("manifest_number") + 1
            img_col = columns.index("image_path") + 1
            roi_col = columns.index("roi_image_path") + 1
            for idx, rec in condensed_df.iterrows():
                r = idx + 2
                # Replace image path columns with filename hyperlinks
                img = rec.get("image_path")
                if img:
                    cell = ws.cell(row=r, column=img_col)
                    fname = Path(img).name
                    cell.value = fname
                    cell.hyperlink = Hyperlink(target=img, display=fname)
                    cell.style = "Hyperlink"
                roi = rec.get("roi_image_path")
                if roi:
                    cell = ws.cell(row=r, column=roi_col)
                    fname = Path(roi).name
                    cell.value = fname
                    cell.hyperlink = Hyperlink(target=roi, display=fname)
                    cell.style = "Hyperlink"
                # Highlight invalid ticket numbers
                if rec.get("ticket_number_valid") != "valid":
                    cell = ws.cell(row=r, column=t_col)
                    cell.fill = red
                    cell.value = rec.get("ticket_number")
                    if roi:
                        cell.hyperlink = Hyperlink(target=roi, display=rec.get("ticket_number"))
                        cell.style = "Hyperlink"
                # Highlight invalid manifest numbers
                if rec.get("manifest_number_valid") != "valid":
                    cell = ws.cell(row=r, column=m_col)
                    cell.fill = red
                    cell.value = rec.get("manifest_number")
                    m_roi = rec.get("manifest_roi_image_path") or roi
                    if m_roi:
                        cell.hyperlink = Hyperlink(target=m_roi, display=rec.get("manifest_number"))
                        cell.style = "Hyperlink"
            wb.save(excel_path)
        except Exception:
            pass

    # Ticket/manifest exception logs
    ticket_exc = df[df["ticket_number"].isna() | (df["ticket_number"] == "")]
    ticket_exc_path = _report_path(cfg, "ticket_number_exceptions_csv", "ticket_number/ticket_number_exceptions.csv")
    if ticket_exc_path:
        os.makedirs(os.path.dirname(ticket_exc_path), exist_ok=True)
        ticket_exc.to_csv(ticket_exc_path, index=False)

    manifest_exc = df[df["manifest_valid"] != "valid"]
    manifest_exc_path = _report_path(cfg, "manifest_number_exceptions_csv", "manifest_number/manifest_number_exceptions.csv")
    if manifest_exc_path:
        os.makedirs(os.path.dirname(manifest_exc_path), exist_ok=True)
        manifest_exc.to_csv(manifest_exc_path, index=False)

    # Summary report
    summary_path = _report_path(cfg, "summary_report", "summary/summary.csv")
    if summary_path:
        os.makedirs(os.path.dirname(summary_path), exist_ok=True)
        summary = {
            "total_pages": len(df),
            "tickets_missing": int(ticket_exc.shape[0]),
            "manifest_valid": int((df["manifest_valid"] == "valid").sum()),
            "manifest_review": int((df["manifest_valid"] == "review").sum()),
            "manifest_invalid": int((df["manifest_valid"] == "invalid").sum()),
        }
        summary_df = pd.DataFrame([summary])

        vendor_counts = (
            df.groupby(["file", "vendor"]).size().reset_index(name="page_count")
        )
        if not vendor_counts.empty:
            vendor_counts["vendor_doc_path"] = vendor_counts.apply(
                lambda r: _make_vendor_doc_path(
                    r["file"], r["vendor"], r["page_count"], cfg
                ),
                axis=1,
            )

        with open(summary_path, "w", newline="", encoding="utf-8") as f:
            summary_df.to_csv(f, index=False)
            if not vendor_counts.empty:
                f.write("\n")
                vendor_counts.to_csv(f, index=False)


def export_preflight_exceptions(exceptions: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    """Write preflight exception rows to CSV if enabled."""
    if not exceptions:
        return
    out_path = _report_path(
        cfg,
        "preflight_exceptions_csv",
        "preflight/preflight_exceptions.csv",
    )
    if out_path:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        pd.DataFrame(exceptions).to_csv(out_path, index=False)


def export_issue_logs(
    ticket_issues: List[Dict[str, Any]],
    issues_log: List[Dict[str, Any]],
    cfg: Dict[str, Any],
) -> None:
    """Write detailed issue logs if enabled."""
    ti_path = _report_path(cfg, "ticket_issues", "ticket_issues.csv")
    if ti_path and ticket_issues:
        os.makedirs(os.path.dirname(ti_path), exist_ok=True)
        pd.DataFrame(ticket_issues).to_csv(ti_path, index=False)

    il_path = _report_path(cfg, "issues_log", "issues_log.csv")
    if il_path and issues_log:
        os.makedirs(os.path.dirname(il_path), exist_ok=True)
        pd.DataFrame(issues_log).to_csv(il_path, index=False)


def export_process_analysis(records: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    """Write per-page OCR timing metrics."""
    path = _report_path(cfg, "process_analysis", "process_analysis.csv")
    if not path or not records:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    pd.DataFrame(records).to_csv(path, index=False)



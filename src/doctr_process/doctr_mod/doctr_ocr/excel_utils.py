import logging
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def color_code_excel(output_dir: str) -> None:
    """Convert ``combined_ticket_numbers.csv`` to XLSX and highlight issues.

    Rows where ``ticket_valid`` is missing or not ``"valid"`` are filled with
    yellow for quick review. If the CSV contains an ``image_path`` column, the
    values in that column are turned into hyperlinks pointing to the image
    files.
    """
    csv_path = os.path.join(output_dir, "combined_ticket_numbers.csv")
    if not os.path.isfile(csv_path):
        logging.warning(f"ticket_numbers.csv not found: {csv_path}")
        return

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

    wb = Workbook()
    ws = wb.active

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    header = rows[0] if rows else []
    image_col = header.index("image_path") + 1 if "image_path" in header else None

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        if image_col and r_idx > 1:
            link_cell = ws.cell(row=r_idx, column=image_col)
            link_cell.hyperlink = link_cell.value

    if "ticket_valid" not in header:
        wb.save(xlsx_path)
        logging.info(f"Excel file saved with highlights: {xlsx_path}")
        return

    status_col = header.index("ticket_valid") + 1
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r_idx in range(2, len(rows) + 1):
        cell = ws.cell(row=r_idx, column=status_col)
        status = str(cell.value or "").lower()
        if status != "valid":
            for c_idx in range(1, len(header) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = highlight

    wb.save(xlsx_path)
    logging.info(f"Excel file saved with highlights: {xlsx_path}")

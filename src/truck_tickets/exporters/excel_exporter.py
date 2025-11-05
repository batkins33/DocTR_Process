"""Excel tracking workbook exporter with 5 sheets matching legacy format.

Generates Excel workbook with the following sheets (Issue #12):
1. All Daily - Combined daily totals across all material types
2. Class2_Daily - Contaminated material by source location
3. Non Contaminated - Clean material tracking
4. Spoils - Waste material by source location (other subs' spoils areas)
5. Import - Materials brought to site

Note: This is the v1.0 scope. Additional sheets (monthly/weekly summaries, cubic yardage)
can be added in future versions.
"""

import logging
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


class ExcelTrackingExporter:
    """Exports truck ticket data to Excel workbook with 5 sheets per spec (Issue #12)."""

    def __init__(self, job_start_date: Optional[date] = None):
        """Initialize Excel exporter.
        
        Args:
            job_start_date: Project start date for job week/month calculations.
                           Defaults to 2024-07-01 for Project 24-105.
        """
        self.logger = logging.getLogger(__name__)
        if openpyxl is None:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        # Default to Project 24-105 start date
        self.job_start_date = job_start_date or date(2024, 7, 1)

    def export(self, tickets: list[dict], output_path: str | Path, job_code: str = None):
        """Export tickets to Excel tracking workbook.

        Args:
            tickets: List of ticket dictionaries
            output_path: Path to output Excel file
            job_code: Optional job code for filtering
        """
        output_path = Path(output_path)
        self.logger.info(f"Generating Excel tracking workbook: {output_path}")

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate 5 required sheets (Issue #12)
        self._create_all_daily_sheet(wb, tickets)
        self._create_class2_daily_sheet(wb, tickets)
        self._create_non_contaminated_sheet(wb, tickets)
        self._create_spoils_sheet(wb, tickets)
        self._create_import_sheet(wb, tickets)

        # Save workbook
        wb.save(output_path)
        self.logger.info(f"✓ Excel workbook saved: {output_path}")

    def _create_all_daily_sheet(self, wb, tickets):
        """Sheet 1: All Daily - Combined daily totals."""
        ws = wb.create_sheet("All Daily")

        # Headers
        headers = ["Date", "Day", "Job Week", "Job Month", "Total", "Class 2",
                   "Non Contaminated", "Spoils", "Notes"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        # Group by date
        daily_data = self._group_by_date(tickets)

        for date_str in sorted(daily_data.keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            day_name = date_obj.strftime("%a")  # Mon, Tue, etc.
            job_week = self._calculate_job_week(date_obj)
            job_month = self._calculate_job_month(date_obj)

            tickets_on_date = daily_data[date_str]
            class2_count = sum(1 for t in tickets_on_date if t.get("material") == "CLASS_2_CONTAMINATED")
            non_contam_count = sum(1 for t in tickets_on_date if t.get("material") == "NON_CONTAMINATED")
            spoils_count = sum(1 for t in tickets_on_date if t.get("material") == "SPOILS")
            total_count = len(tickets_on_date)

            ws.append([
                date_str,
                day_name,
                job_week,
                job_month,
                total_count,
                class2_count,
                non_contam_count,
                spoils_count,
                ""  # Notes
            ])

    def _create_class2_daily_sheet(self, wb, tickets):
        """Sheet 2: Class2_Daily - Contaminated material by source location."""
        ws = wb.create_sheet("Class2_Daily")

        # Headers - source locations
        headers = ["Date", "Day", "Job Week", "Job Month", "Total",
                   "PODIUM", "Zone E GARAGE", "South_MSE Wall", "MSE Wall",
                   "PierEx", "Pond", "South Fill", "Tract 2", "Notes"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        # Filter contaminated tickets
        class2_tickets = [t for t in tickets if t.get("material") == "CLASS_2_CONTAMINATED"]
        daily_data = self._group_by_date(class2_tickets)

        for date_str in sorted(daily_data.keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            tickets_on_date = daily_data[date_str]

            # Count by source
            source_counts = defaultdict(int)
            for ticket in tickets_on_date:
                source = ticket.get("source", "UNKNOWN")
                source_counts[source] += 1

            ws.append([
                date_str,
                date_obj.strftime("%a"),
                self._calculate_job_week(date_obj),
                self._calculate_job_month(date_obj),
                len(tickets_on_date),
                source_counts.get("PODIUM", 0),
                source_counts.get("ZONE_E_GARAGE", 0),
                source_counts.get("SOUTH_MSE_WALL", 0),
                source_counts.get("MSE_WALL", 0),
                source_counts.get("PIER_EX", 0),
                source_counts.get("POND", 0),
                source_counts.get("SOUTH_FILL", 0),
                source_counts.get("TRACT_2", 0),
                ""
            ])

    def _create_non_contaminated_sheet(self, wb, tickets):
        """Sheet 3: Non Contaminated - Clean material tracking."""
        ws = wb.create_sheet("Non Contaminated")

        headers = ["Date", "Day", "Job Week", "Job Month", "Total", "SPG", "Spoils", "Location"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        # Filter non-contaminated tickets
        non_contam_tickets = [t for t in tickets if t.get("material") == "NON_CONTAMINATED"]
        daily_data = self._group_by_date(non_contam_tickets)

        for date_str in sorted(daily_data.keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            tickets_on_date = daily_data[date_str]

            spg_count = sum(1 for t in tickets_on_date if t.get("source") == "SPG")
            destinations = set(t.get("destination", "UNKNOWN") for t in tickets_on_date)

            ws.append([
                date_str,
                date_obj.strftime("%a"),
                self._calculate_job_week(date_obj),
                self._calculate_job_month(date_obj),
                len(tickets_on_date),
                spg_count,
                0,  # Spoils column
                ", ".join(destinations)
            ])

    def _create_spoils_sheet(self, wb, tickets):
        """Sheet 4: Spoils - Waste material by SOURCE location (other subs' spoils areas).
        
        IMPORTANT: Per spec clarification, spoils locations are SOURCES (other subs' staging areas),
        not destinations. All spoils go to WM Lewisville destination.
        """
        ws = wb.create_sheet("Spoils")

        headers = ["Date", "Day", "Job Week", "Job Month", "Total",
                   "BECK SPOILS", "NTX Spoils", "UTX Spoils", "MVP-TC1", "MVP-TC2", "Notes"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        # Filter spoils tickets
        spoils_tickets = [t for t in tickets if t.get("material") == "SPOILS"]
        daily_data = self._group_by_date(spoils_tickets)

        for date_str in sorted(daily_data.keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            tickets_on_date = daily_data[date_str]

            # Count by SOURCE (not destination - spoils sources are other subs' areas)
            source_counts = defaultdict(int)
            for ticket in tickets_on_date:
                source = ticket.get("source", "UNKNOWN")
                source_counts[source] += 1

            ws.append([
                date_str,
                date_obj.strftime("%a"),
                self._calculate_job_week(date_obj),
                self._calculate_job_month(date_obj),
                len(tickets_on_date),
                source_counts.get("BECK_SPOILS", 0),
                source_counts.get("NTX_SPOILS", 0),
                source_counts.get("UTX_SPOILS", 0),
                source_counts.get("MVP_TC1", 0),
                source_counts.get("MVP_TC2", 0),
                ""
            ])

    def _create_import_sheet(self, wb, tickets):
        """Sheet 5: Import - Materials brought to site (9 material types)."""
        ws = wb.create_sheet("Import")

        headers = ["DATE", "3X5", "ASPHALT", "C2", "DIRT", "FILL", "FLEX",
                   "FLEXBASE", "ROCK", "UTILITY STONE", "Grand Total"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        # Filter import tickets
        import_tickets = [t for t in tickets if t.get("ticket_type") == "IMPORT"]
        daily_data = self._group_by_date(import_tickets)

        for date_str in sorted(daily_data.keys()):
            tickets_on_date = daily_data[date_str]

            # Count by material type
            material_counts = defaultdict(int)
            for ticket in tickets_on_date:
                material = ticket.get("material", "UNKNOWN")
                material_counts[material] += 1

            ws.append([
                date_str,
                material_counts.get("3X5", 0),
                material_counts.get("ASPHALT", 0),
                material_counts.get("C2", 0),
                material_counts.get("DIRT", 0),
                material_counts.get("FILL", 0),
                material_counts.get("FLEX", 0),
                material_counts.get("FLEXBASE", 0),
                material_counts.get("ROCK", 0),
                material_counts.get("UTILITY_STONE", 0),
                len(tickets_on_date)
            ])

    # Helper methods

    def _group_by_date(self, tickets):
        """Group tickets by date."""
        daily_data = defaultdict(list)
        for ticket in tickets:
            date_str = ticket.get("ticket_date", "UNKNOWN")
            daily_data[date_str].append(ticket)
        return daily_data

    def _calculate_job_week(self, date_obj):
        """Calculate job week string (e.g., 'Week 16 - (End 10/20/24)')."""
        from ..utils.date_calculations import calculate_job_week
        # Convert datetime to date if needed
        if isinstance(date_obj, datetime):
            date_obj = date_obj.date()
        return calculate_job_week(date_obj, self.job_start_date)

    def _calculate_job_month(self, date_obj):
        """Calculate job month string (e.g., '004 - October 24')."""
        from ..utils.date_calculations import calculate_job_month
        # Convert datetime to date if needed
        if isinstance(date_obj, datetime):
            date_obj = date_obj.date()
        return calculate_job_month(date_obj, self.job_start_date)

    def _style_header_row(self, ws, row_num):
        """Apply styling to header row."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
